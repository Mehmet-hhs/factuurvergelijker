# reporter.py
# Fase 3: Excel export en rapportage

"""
reporter.py
===========

Verantwoordelijkheid:
    Resultaten presenteren in bruikbare formaten voor business.

Functies:
    - genereer_samenvatting: Berekent metrics per status
    - exporteer_naar_excel: Genereert Excel met 2 tabbladen (samenvatting + details)

Output:
    Excel-bestand met kleurcodering, autofilters en leesbare samenvattingen.
"""

import pandas as pd
from pathlib import Path
from typing import Dict
from datetime import datetime
import sys

# Voeg parent directory toe zodat config.py gevonden kan worden
sys.path.append(str(Path(__file__).parent.parent))
import config

# Import voor Excel styling
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


def _schrijf_details_sheet(worksheet, df_resultaat: pd.DataFrame):
    """
    Schrijft detailgegevens naar Excel sheet met opmaak.
    """
    
    # ✨ DEBUG: Print aantal rijen
    print(f"📊 _schrijf_details_sheet: {len(df_resultaat)} rijen te schrijven")
    
    # Schrijf DataFrame naar sheet
    for r_idx, row in enumerate(dataframe_to_rows(df_resultaat, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
            
            # Header styling
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
    
    # ✨ DEBUG: Print status kolom index
    status_col_idx = df_resultaat.columns.get_loc('status') + 1
    print(f"🎨 Status kolom index: {status_col_idx}")
    
    # Kleurcodering voor status kolom
    aantal_gekleurd = 0
    for rij_idx in range(2, len(df_resultaat) + 2):  # Start na header
        status_cell = worksheet.cell(row=rij_idx, column=status_col_idx)
        status_waarde = status_cell.value
        
        # ✨ DEBUG: Print eerste 5 statussen
        if rij_idx <= 6:
            print(f"   Rij {rij_idx}: status = '{status_waarde}'")
        
        if status_waarde == config.STATUS_OK:
            status_cell.fill = _get_fill_color('green')
            aantal_gekleurd += 1
        elif status_waarde == config.STATUS_AFWIJKING:
            status_cell.fill = _get_fill_color('orange')
            aantal_gekleurd += 1
        elif status_waarde in [config.STATUS_ONTBREEKT_FACTUUR, config.STATUS_ONTBREEKT_SYSTEEM]:
            status_cell.fill = _get_fill_color('red')
            aantal_gekleurd += 1
        elif status_waarde == config.STATUS_GEDEELTELIJK:
            status_cell.fill = _get_fill_color('yellow')
            aantal_gekleurd += 1
        elif status_waarde == config.STATUS_FOUT:
            status_cell.fill = _get_fill_color('gray')
            aantal_gekleurd += 1
    
    # ✨ DEBUG: Print totaal gekleurd
    print(f"✅ Aantal cellen gekleurd: {aantal_gekleurd}/{len(df_resultaat)}")
    
    # Autofilter toevoegen
    worksheet.auto_filter.ref = worksheet.dimensions
    
    # ... rest van de code ...

def genereer_samenvatting(df_resultaat: pd.DataFrame) -> Dict:
    """
    Genereert samenvattende statistieken van het vergelijkingsresultaat.
    
    Parameters
    ----------
    df_resultaat : pd.DataFrame
        Resultaat-DataFrame van comparator.py.
    
    Returns
    -------
    dict
        {
            'totaal_regels': int,
            'status_counts': {
                'OK': int,
                'AFWIJKING': int,
                'ONTBREEKT OP FACTUUR': int,
                'ONTBREEKT IN SYSTEEM': int,
                'GEDEELTELIJK': int,
                'FOUT': int
            }
        }
    
    Voorbeelden
    -----------
    >>> samenvatting = genereer_samenvatting(resultaat_df)
    >>> print(samenvatting['totaal_regels'])
    247
    >>> print(samenvatting['status_counts']['OK'])
    198
    """
    
    # Tel totaal aantal regels
    totaal_regels = len(df_resultaat)
    
    # Tel per status
    status_counts = df_resultaat['status'].value_counts().to_dict()
    
    # Vul ontbrekende statussen aan met 0
    alle_statussen = [
        config.STATUS_OK,
        config.STATUS_AFWIJKING,
        config.STATUS_ONTBREEKT_FACTUUR,
        config.STATUS_ONTBREEKT_SYSTEEM,
        config.STATUS_GEDEELTELIJK,
        config.STATUS_FOUT
    ]
    
    for status in alle_statussen:
        if status not in status_counts:
            status_counts[status] = 0
    
    samenvatting = {
        'totaal_regels': totaal_regels,
        'status_counts': status_counts
    }
    
    return samenvatting


def exporteer_naar_excel(
    df_resultaat: pd.DataFrame,
    output_pad: Path,
    bestandsnaam_systeem: str = "systeem",
    bestandsnaam_factuur: str = "factuur"
) -> Path:
    """
    Exporteert resultaten naar Excel met 2 tabbladen en kleurcodering.
    
    Parameters
    ----------
    df_resultaat : pd.DataFrame
        Resultaat-DataFrame van comparator.py.
    output_pad : Path
        Directory waar Excel-bestand opgeslagen wordt.
    bestandsnaam_systeem : str
        Naam van systeemexport (voor in bestandsnaam).
    bestandsnaam_factuur : str
        Naam van leveranciersfactuur (voor in bestandsnaam).
    
    Returns
    -------
    Path
        Volledig pad naar gegenereerd Excel-bestand.
    
    Voorbeelden
    -----------
    >>> pad = exporteer_naar_excel(
    ...     resultaat_df,
    ...     Path("./output"),
    ...     "export_jan",
    ...     "factuur_leverancier_A"
    ... )
    >>> print(pad)
    ./output/vergelijking_export_jan_vs_factuur_leverancier_A_20240215_143022.xlsx
    """
    
    # Genereer bestandsnaam met timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    bestandsnaam = f"vergelijking_{bestandsnaam_systeem}_vs_{bestandsnaam_factuur}_{timestamp}.xlsx"
    
    volledig_pad = output_pad / bestandsnaam
    
    # Zorg dat output directory bestaat
    output_pad.mkdir(parents=True, exist_ok=True)
    
    # Genereer samenvatting
    samenvatting = genereer_samenvatting(df_resultaat)
    
    # Maak Excel workbook
    workbook = Workbook()
    
    # Verwijder standaard sheet
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']
    
    # TABBLAD 1: Samenvatting
    ws_samenvatting = workbook.create_sheet(config.EXCEL_SHEET_SAMENVATTING)
    _schrijf_samenvatting_sheet(ws_samenvatting, samenvatting)
    
    # TABBLAD 2: Details
    ws_details = workbook.create_sheet(config.EXCEL_SHEET_NAAM)
    _schrijf_details_sheet(ws_details, df_resultaat)
    
    # Sla op
    workbook.save(volledig_pad)
    
    return volledig_pad


def _schrijf_samenvatting_sheet(worksheet, samenvatting: Dict):
    """
    Schrijft samenvattingsgegevens naar Excel sheet.
    
    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        Excel worksheet om naar te schrijven.
    samenvatting : dict
        Samenvatting zoals gegenereerd door genereer_samenvatting().
    """
    
    # Titel
    worksheet['A1'] = 'VERGELIJKINGSRESULTAAT SAMENVATTING'
    worksheet['A1'].font = Font(bold=True, size=14)
    
    # Totaal regels
    worksheet['A3'] = 'Totaal regels verwerkt:'
    worksheet['B3'] = samenvatting['totaal_regels']
    worksheet['A3'].font = Font(bold=True)
    
    # Status breakdown
    worksheet['A5'] = 'Status'
    worksheet['B5'] = 'Aantal'
    worksheet['A5'].font = Font(bold=True)
    worksheet['B5'].font = Font(bold=True)
    
    status_counts = samenvatting['status_counts']
    
    rij = 6
    status_mapping = [
        (config.STATUS_OK, '✅ OK', 'green'),
        (config.STATUS_AFWIJKING, '⚠️ Afwijking', 'orange'),
        (config.STATUS_ONTBREEKT_FACTUUR, '❌ Ontbreekt op factuur', 'red'),
        (config.STATUS_ONTBREEKT_SYSTEEM, '❌ Ontbreekt in systeem', 'red'),
        (config.STATUS_GEDEELTELIJK, '⚡ Gedeeltelijk', 'yellow'),
        (config.STATUS_FOUT, '⛔ Fout', 'gray')
    ]
    
    for status_key, status_label, kleur in status_mapping:
        worksheet[f'A{rij}'] = status_label
        worksheet[f'B{rij}'] = status_counts.get(status_key, 0)
        
        # Kleurcodering
        fill_color = _get_fill_color(kleur)
        worksheet[f'A{rij}'].fill = fill_color
        
        rij += 1
    
    # Kolombreedte aanpassen
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 15


def _schrijf_details_sheet(worksheet, df_resultaat: pd.DataFrame):
    """
    Schrijft detailgegevens naar Excel sheet met opmaak.
    
    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        Excel worksheet om naar te schrijven.
    df_resultaat : pd.DataFrame
        Resultaat-DataFrame met alle details.
    """
    
    # Schrijf DataFrame naar sheet
    for r_idx, row in enumerate(dataframe_to_rows(df_resultaat, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
            
            # Header styling
            if r_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
    
    # Kleurcodering voor status kolom (kolom A)
    status_col_idx = df_resultaat.columns.get_loc('status') + 1
    
    for rij_idx in range(2, len(df_resultaat) + 2):  # Start na header
        status_cell = worksheet.cell(row=rij_idx, column=status_col_idx)
        status_waarde = status_cell.value
        
        if status_waarde == config.STATUS_OK:
            status_cell.fill = _get_fill_color('green')
        elif status_waarde == config.STATUS_AFWIJKING:
            status_cell.fill = _get_fill_color('orange')
        elif status_waarde in [config.STATUS_ONTBREEKT_FACTUUR, config.STATUS_ONTBREEKT_SYSTEEM]:
            status_cell.fill = _get_fill_color('red')
        elif status_waarde == config.STATUS_GEDEELTELIJK:
            status_cell.fill = _get_fill_color('yellow')
        elif status_waarde == config.STATUS_FOUT:
            status_cell.fill = _get_fill_color('gray')
    
    # Autofilter toevoegen
    worksheet.auto_filter.ref = worksheet.dimensions
    
    # Kolombreedte automatisch aanpassen
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Max 50 voor leesbaarheid
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _get_fill_color(kleur_naam: str) -> PatternFill:
    """
    Helpt functie: retourneert PatternFill voor gegeven kleurnaam.
    
    Parameters
    ----------
    kleur_naam : str
        'green', 'orange', 'red', 'yellow', of 'gray'.
    
    Returns
    -------
    PatternFill
        Excel fill object.
    """
    
    kleuren = {
        'green': 'C6EFCE',
        'orange': 'FFCC99',
        'red': 'FFC7CE',
        'yellow': 'FFEB9C',
        'gray': 'D9D9D9'
    }
    
    hex_color = kleuren.get(kleur_naam, 'FFFFFF')
    
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")